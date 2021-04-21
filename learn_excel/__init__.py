# from openpyxl import Workbook
# workbook = Workbook()
# sheet = workbook.active
# workbook.save(filename="bdika.xlsx")


from openpyxl import *
xl=load_workbook("bdika_2.xlsx")
ss=xl.active
def p():
    for i in ss.iter_rows(values_only=True):
        print(i)

# ss["F1"]="hey"
p()